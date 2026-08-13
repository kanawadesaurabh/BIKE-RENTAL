from django.shortcuts import render
from django.utils import timezone
from django.db.models import Sum
from decimal import Decimal
from datetime import timedelta
from django.utils import timezone
from bikes.models import Bike
from customers.models import Customer
from rentals.models import Rental
from payments.models import Payment
from expenses.models import Expense

from django.http import HttpResponse
from openpyxl import Workbook


# =====================================================
# DASHBOARD
# =====================================================


def excel_datetime(value):

    if value is not None and timezone.is_aware(value):
        return timezone.make_naive(value)

    return value



def export_excel(request):

    wb = Workbook()

    # ==================================================
    # 1. BIKES SHEET
    # ==================================================

    ws = wb.active
    ws.title = "Bikes"

    ws.append([
        "ID",
        "Bike Name",
        "Brand",
        "Model",
        "Registration Number",
        "Color",
        "Year",
        "Daily Rent",
        "Security Deposit",
        "Status"
    ])

    for bike in Bike.objects.all():

        ws.append([
            bike.id,
            bike.bike_name,
            bike.brand,
            bike.model,
            bike.registration_number,
            bike.color,
            bike.year,
            bike.daily_rent,
            bike.security_deposit,
            bike.status
        ])

    # ==================================================
    # 2. CUSTOMERS SHEET
    # ==================================================

    ws = wb.create_sheet("Customers")

    ws.append([
        "ID",
        "Customer Name",
        "Mobile",
        "Aadhaar Number",
        "Driving License",
        "Address"
    ])

    for customer in Customer.objects.all():

        ws.append([
            customer.id,
            customer.customer_name,
            customer.mobile,
            customer.aadhaar_number,
            customer.driving_license,
            customer.address
        ])

    # ==================================================
    # 3. RENTALS SHEET
    # ==================================================

    ws = wb.create_sheet("Rentals")

    ws.append([
        "ID",
        "Customer",
        "Mobile",
        "Bike",
        "Registration No",
        "Rental Type",
        "Rental Days",
        "Rent Date",
        "Expected Return",
        "Actual Return",
        "Daily Rent",
        "Security Deposit",
        "Advance Payment",
        "Total Days",
        "Total Amount",
        "Late Fine",
        "Damage Charge",
        "Manual Extra Charge",
        "Deposit Refund",
        "Remaining Amount",
        "Remarks",
        "Status"
    ])

    rentals = Rental.objects.select_related(
        "customer",
        "bike"
    ).all()

    for rental in rentals:

        ws.append([
            rental.id,
            rental.customer.customer_name,
            rental.customer.mobile,
            rental.bike.bike_name,
            rental.bike.registration_number,
            rental.rental_type,
            rental.rental_days,
excel_datetime(rental.rent_date),
excel_datetime(rental.expected_return_date),
excel_datetime(rental.actual_return_date),
            rental.daily_rent,
            rental.security_deposit,
            rental.advance_payment,
            rental.total_days,
            rental.total_amount,
            rental.late_fine,
            rental.damage_charge,
            rental.manual_extra_charge,
            rental.deposit_refund,
            rental.remaining_amount,
            rental.remarks,
            rental.status
        ])

    # ==================================================
    # 4. PAYMENTS SHEET
    # ==================================================

    ws = wb.create_sheet("Payments")

    ws.append([
        "ID",
        "Customer",
        "Bike",
        "Registration No",
        "Payment Date",
        "Amount",
        "Payment Mode",
        "Remarks"
    ])

    payments = Payment.objects.select_related(
        "rental__customer",
        "rental__bike"
    ).all()

    for payment in payments:

        ws.append([
            payment.id,
            payment.rental.customer.customer_name,
            payment.rental.bike.bike_name,
            payment.rental.bike.registration_number,
            payment.payment_date,
            payment.amount,
            payment.payment_mode,
            payment.remarks
        ])

    # ==================================================
    # AUTO COLUMN WIDTH
    # ==================================================

    for ws in wb.worksheets:

        for column in ws.columns:

            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:

                if cell.value is not None:

                    cell_length = len(str(cell.value))

                    if cell_length > max_length:
                        max_length = cell_length

            ws.column_dimensions[column_letter].width = min(
                max_length + 2,
                40
            )

        # Freeze first row
        ws.freeze_panes = "A2"

    # ==================================================
    # DOWNLOAD EXCEL FILE
    # ==================================================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="BikeRental_Report.xlsx"'
    )

    wb.save(response)

    return response
def dashboard(request):

    # =========================
    # TODAY
    # =========================

    today = timezone.localdate()

    # =========================
    # TOMORROW
    # =========================

    tomorrow = today + timedelta(days=1)

    # =========================
    # BIKE SUMMARY
    # =========================

    total_bikes = Bike.objects.count()

    available_bikes = Bike.objects.filter(
        status="Available"
    ).count()

    rented_bikes = Bike.objects.filter(
        status="Rented"
    ).count()

    # =========================
    # CUSTOMER SUMMARY
    # =========================

    total_customers = Customer.objects.count()

    # =========================
    # ACTIVE RENTALS
    # =========================

    active_rentals = Rental.objects.filter(
        status="Active"
    ).count()

    # =========================
    # RETURN TODAY
    # =========================

    return_today = Rental.objects.filter(
        status="Active",
        expected_return_date__date=today
    ).select_related(
        "customer",
        "bike"
    ).order_by(
        "expected_return_date"
    )

    # =========================
    # OVERDUE RENTALS
    # =========================

    overdue_rentals = Rental.objects.filter(
        status="Active",
        expected_return_date__date__lt=today
    ).select_related(
        "customer",
        "bike"
    ).order_by(
        "expected_return_date"
    )

    # =========================
    # UPCOMING RETURNS
    # =========================

    upcoming_returns = Rental.objects.filter(
        status="Active",
        expected_return_date__date=tomorrow
    ).select_related(
        "customer",
        "bike"
    ).order_by(
        "expected_return_date"
    )

    # =========================
    # PENDING PAYMENTS
    # =========================

    pending_rentals = Rental.objects.filter(
        status="Active",
        remaining_amount__gt=0
    ).select_related(
        "customer",
        "bike"
    ).order_by(
        "-remaining_amount"
    )

    pending_payments = pending_rentals.count()

    # =========================
    # TODAY INCOME
    # =========================

    today_income = Payment.objects.filter(
        payment_date=today
    ).aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")

    # =========================
    # MONTHLY INCOME
    # =========================

    current_month = today.month
    current_year = today.year

    monthly_income = Payment.objects.filter(
        payment_date__month=current_month,
        payment_date__year=current_year
    ).aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")

    # =========================
    # TOTAL INCOME
    # =========================

    total_income = Payment.objects.aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")

    # =========================
    # TOTAL EXPENSE
    # =========================

    total_expense = Expense.objects.aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")

    # =========================
    # TODAY EXPENSE
    # =========================

    today_expense = Expense.objects.filter(
        expense_date=today
    ).aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")

    # =========================
    # TOTAL PROFIT
    # =========================

    total_profit = total_income - total_expense

    total_profit = total_profit.quantize(
        Decimal("0.01")
    )

    # =========================
    # RECENT PAYMENTS
    # =========================

    recent_payments = Payment.objects.select_related(
        "rental",
        "rental__customer",
        "rental__bike"
    ).order_by(
        "-payment_date"
    )[:5]

    # =========================
    # RECENT RENTALS
    # =========================

    recent_rentals = Rental.objects.select_related(
        "customer",
        "bike"
    ).order_by(
        "-rent_date"
    )[:5]

    # =========================
    # RECENT EXPENSES
    # =========================

    recent_expenses = Expense.objects.order_by(
        "-expense_date"
    )[:5]

    # =========================
    # DASHBOARD CONTEXT
    # =========================

    context = {
        "today": today,

        # Bikes
        "total_bikes": total_bikes,
        "available_bikes": available_bikes,
        "rented_bikes": rented_bikes,

        # Customers
        "total_customers": total_customers,

        # Rentals
        "active_rentals": active_rentals,

        # Alerts
        "return_today": return_today,
        "overdue_rentals": overdue_rentals,
        "upcoming_returns": upcoming_returns,

        # Payments
        "pending_rentals": pending_rentals,
        "pending_payments": pending_payments,

        # Income
        "today_income": today_income,
        "monthly_income": monthly_income,
        "total_income": total_income,

        # Expenses
        "today_expense": today_expense,
        "total_expense": total_expense,

        # Profit
        "total_profit": total_profit,

        # Recent Activity
        "recent_payments": recent_payments,
        "recent_rentals": recent_rentals,
        "recent_expenses": recent_expenses,
    }

    return render(
        request,
        "dashboard/dashboard.html",
        context
    )


# =====================================================
# BIKE SEARCH
# =====================================================

def bike_search(request):

    registration_number = request.GET.get(
        "registration_number"
    )

    bike = None
    rental = None

    if registration_number:

        bike = Bike.objects.filter(
            registration_number__iexact=registration_number
        ).first()

        if bike:

            rental = Rental.objects.filter(
                bike=bike,
                status="Active"
            ).select_related(
                "customer",
                "bike"
            ).first()

    return render(
        request,
        "dashboard/bike_search.html",
        {
            "bike": bike,
            "rental": rental,
            "registration_number": registration_number,
            "today": timezone.localdate(),
        }
    )